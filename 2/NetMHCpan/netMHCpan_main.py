from selenium import webdriver
from openpyxl import load_workbook
from NetMHCpan.Modules import SalvarEstado
import os
import pyautogui
import clipboard

diretorio = "C:\\Users\\mauro\\Desktop\\Bruno\\Automação\\SCRIPTS\\NetMHCpan\\"
driver_path = "C:\\Users\\mauro\\Desktop\\Bruno\\Automação\\configuring the envyronmet\\chromedriver\\chromedriver_win32\\chromedriver.exe"

#variaveis
ultima_analise = 1
numero_analises = 0
n_alelos_remanescentes = 0
alelos = []
linha = int
final = bool

alelos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\alelostestefinal.xlsx"
book_alelos = load_workbook(alelos_path)
sheet = book_alelos.active

epitopos_path = str
epitopos_txt = str
epitopos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\Epitopos\\"
epitopos = ""

#Tenta abrir o arquivo dos epitopos
abrir = False
while not abrir:
    try:
        nome_epitopos = str(input("Entre com o nome do epitopo sem a extensão .txt:"))
        epitopos_txt = open(epitopos_path + nome_epitopos+'.txt', "r")
        for linha in epitopos_txt:
            if not epitopos == "":
                epitopos += '\n' + linha.replace('\n', "")
            else:
                epitopos = linha

        abrir = True

    except:
        print("Nome incorreto. Insira o nome sem a extensão .txt")


# checa se há registros
try:
    file = open(diretorio + nome_epitopos + '_estado.txt', 'r')
    estados = file.readlines()
    print(estados)
    file.close() #TODO: testar estados[0], [1] e [2] e os tipos
    ultima_analise = int(estados[0])
    numero_analises = int(estados[1])
    n_alelos_remanescentes = int(estados[2]) #TODO: testar essas 3 variaveis do estado
    print("Existe um estado")
    print("ultima analise: ", ultima_analise, type(ultima_analise))
    print("numero_analises: ", numero_analises, type(numero_analises))
    print("n_alelos_remanescentes: ", n_alelos_remanescentes, type(n_alelos_remanescentes))

    # checa se a ultima_analise feita é a análise final
    if ultima_analise == int(estados[1]):
        print("Analise Final")
        final = True
    else:
        final = False

except:
    print("Não há estados")
    final = False
    linha = 0
    while (not final):
        linha = linha + 1
        print(sheet["A"+str(linha)].value)
        if str(sheet["A" + str(linha)].value) == "None":
            print("Acabou")
            final = True
            linha = linha - 1
    final = False
    print("numero de alelos: ", linha)
    print("numero de analises de 10: ", int(linha / 10))
    print("numero de alelos da ultima analise: ", linha % 10)

    #checa se todas as análises serão de 10 alelos
    if (linha % 10 == 0):
        numero_analises = int(linha / 10)

    #acrescenta uma analise a mais para os alelos remanescentes
    else:
        numero_analises = int(linha / 10) + 1
        n_alelos_remanescentes = linha % 10

    #cria o estado
    # ultima analise e numeroamalises estao com valores trocados
    file = open(diretorio + nome_epitopos + '_estado.txt', 'w')
    file.write(str(ultima_analise) + '\n' + str(numero_analises) + '\n' + str(n_alelos_remanescentes))
    file.close()

    # checa se há registros
    file = open(diretorio + nome_epitopos + '_estado.txt', 'r')
    estados = file.readlines()
    print(estados)
    estado = file.readline()
    print(estado)
    file.close()

while(ultima_analise != numero_analises):

    #netMHCpan
    print("Entra no netMHCpan e coloca os epitopos")
    driver = webdriver.Chrome(driver_path)
    driver.get("http://www.cbs.dtu.dk/services/NetMHCpan/")
    driver.find_element_by_name("SEQPASTE").click()

    #É aqui
    clipboard.copy(epitopos)
    pyautogui.hotkey('ctrl', 'v')

    print("pega os alelos")
    # alelos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\alelostestefinal.xlsx"
    # book_alelos = load_workbook(alelos_path)
    # sheet = book_alelos.active
    print("Reseta a variável dos alelos para pegar os alelos da próxima análise")
    alelos = []
    print("Insere os seguintes alelos:")

    if (not final):
        print("insere os 10")
        for i in range(10):
            print(i+1, sheet["A"+str((10*(ultima_analise - 1)) + 1 + i)].value)
            if not i == 9:
                alelos.append(sheet["A"+str((10*(ultima_analise - 1)) + 1 + i)].value + ",")
            else:
                alelos.append(sheet["A"+str((10*(ultima_analise - 1)) + 1 + i)].value)
        print(alelos)

    #netMHCpan
    print("netMHCpan: coloca os alelos")
    alele_area = driver.find_element_by_name("allele")
    alele_area.send_keys(alelos)

    driver.find_element_by_name("BApred").click()
    driver.find_element_by_name("sort").click()
    driver.find_element_by_name("xlsdump").click()

    print("netMHCpan: submete p/ analise e gera a saída")
    # Submit
    inputList = driver.find_elements_by_tag_name("input")
    inputList[-2].click()

    print("nethMHCpan: gera o excel de saida")
    # Output
    clicou = False
    esperando = True
    while (not clicou):
        try:
            driver.find_elements_by_tag_name("a")[-3].click()
            clicou = True
        except:
            if esperando:
                print("Esperando carregar")
                esperando = False

    print("salvando o estado")
    ultima_analise = ultima_analise + 1
    SalvarEstado.salvarEstado(diretorio, nome_epitopos, numero_analises, ultima_analise, n_alelos_remanescentes)
else:
    # netMHCpan
    print("Entra no netMHCpan e coloca os epitopos")
    driver = webdriver.Chrome(driver_path)
    driver.get("http://www.cbs.dtu.dk/services/NetMHCpan/")
    textarea = driver.find_element_by_tag_name('textarea')
    textarea.send_keys(epitopos)

    alelos = []
    print("insere os remanescentes")
    for i in range(n_alelos_remanescentes):
        print(i+1, sheet["A"+str((10*(ultima_analise - 1)) + 1 + i)].value)
        if not i == (n_alelos_remanescentes-1):
            alelos.append(sheet["A"+str((10*(ultima_analise - 1)) + 1 + i)].value + ",")
        else:
            alelos.append(sheet["A"+str((10*(ultima_analise - 1)) + 1 + i)].value)
    print(alelos)

    # netMHCpan
    print("netMHCpan: coloca os alelos")
    alele_area = driver.find_element_by_name("allele")
    alele_area.send_keys(alelos)

    driver.find_element_by_name("BApred").click()
    driver.find_element_by_name("sort").click()
    driver.find_element_by_name("xlsdump").click()

    print("netMHCpan: submete p/ analise e gera a saída")
    # Submit
    inputList = driver.find_elements_by_tag_name("input")
    inputList[-2].click()

    print("nethMHCpan: gera o excel de saida")
    # Output
    clicou = False
    esperando = True
    while (not clicou):
        try:
            driver.find_elements_by_tag_name("a")[-3].click()
            clicou = True
        except:
            if esperando:
                print("Esperando carregar")
                esperando = False


    print("apagando o estado")
    os.remove(diretorio + nome_epitopos + '_estado.txt')

print("##############Finalizado################")









