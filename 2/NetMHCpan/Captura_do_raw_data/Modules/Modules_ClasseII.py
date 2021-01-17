global epitopo
global results
global excel_path

def putResultsIntoExcel(epitopo, results, excel_path):

    from openpyxl import load_workbook

    try:
        book_result = load_workbook(excel_path + str(epitopo) + "_results_HLAs_rankingMinimo_binders.xlsx")
        print("acessando o arquivo excel existente")

    except:
        print('criando o arquivo excel')

        # filename = nome_fasta, write_only = True
        book_modelo = load_workbook(excel_path + "result_modelo.xlsx")
        book_result = book_modelo

        # book.create_sheet()
        book_result.save(excel_path + str(epitopo) + "_results_HLAs_rankingMinimo_binders.xlsx")
        book_modelo.close()

    # parametrizar: abrindo o arquivo excel (já existente ou criado)
    sheet = book_result.active

    linha_vazia = False
    linha_inserção = int
    i = 2
    while not linha_vazia:

        if str(sheet["A" + str(i)].value) == "None":
            linha_vazia = True
            linha_inserção = i
        else:
            i = i + 1


    for index, elem in enumerate(results):
        print(elem)
        if index == 0:
            sheet["A" + str(linha_inserção)] = elem
        elif index == 1:
            sheet["B" + str(linha_inserção)] = elem
        elif index == 2:
            sheet["C" + str(linha_inserção)] = elem
        elif index == 3:
            sheet["D" + str(linha_inserção)] = elem
        elif index == 4:
            sheet["E" + str(linha_inserção)] = elem

    book_result.save(excel_path + str(epitopo) + "_results_HLAs_rankingMinimo_binders.xlsx")

    print('fechando o arquivo excel')
    book_result.close() # nao vai pra main, pois a funcao do excel já vai receber o arquivo do excel aberto
# excel_path = "C:\\Users\\Bruno Conde\\Documents\\LGHM\\Automação\\SCRIPTS\\netMHCpan\\results_netMHCpan\\"
# epitopo = "VarVirus"
#
# putResultsIntoExcel(epitopo, ["hla x", 3, 4, 9], excel_path)

global excel_path
global sheet

def getHLAs(sheet):
    hla_list = [] # retorna uma lista de hlas do arquivo excel da analise

    if sheet["E1"] != 'None':
        hla_list.append(sheet["E1"].value)
        if sheet["H1"] != 'None':
            hla_list.append(sheet["H1"].value)
            if sheet["K1"] != 'None':
                hla_list.append(sheet["K1"].value)
                if sheet["N1"] != 'None':
                    hla_list.append(sheet["N1"].value)
                    if sheet["Q1"] != 'None':
                        hla_list.append(sheet["Q1"].value)
                        if sheet["T1"] != 'None':
                            hla_list.append(sheet["T1"].value)
                            if sheet["W1"] != 'None':
                                hla_list.append(sheet["W1"].value)
                                if sheet["Z1"] != 'None':
                                    hla_list.append(sheet["Z1"].value)
                                    if sheet["AC1"] != 'None':
                                        hla_list.append(sheet["AC1"].value)
                                        if sheet["AF1"] != 'None':
                                            hla_list.append(sheet["AF1"].value)

    print("n_HLAs: ", len(hla_list))
    print(hla_list)

    return hla_list
# excel_path = "C:\\Users\\Bruno Conde\\Documents\\LGHM\\Automação\\SCRIPTS\\netMHCpan\\results_netMHCpan\\"
# nome_excel = "teste"
# book = load_workbook(excel_path + nome_excel + ".xlsx")
# sheet = book.active
# getHLAs(sheet)

global n_remanescentes

def getEachExcel(n_remanescentes):
    abriu = False
    while not abriu:
        try:
            n_excelFiles = int(input("Entre com o numero de arquivos excel:" + '\n'))
            abriu = True
            print("n_excelFiles: ", n_excelFiles)
        except:
            print("Entre com um número inteiro")

    for index in range (n_excelFiles):
        print("Abrindo o arquivo " + str(index + 1))

        if not (index + 1) == n_excelFiles:
            print("Obtendo o resultdado dos hlaS de 10 ")
        elif n_remanescentes != 0:
            print("Obtendo resuldado dos "+ str(n_remanescentes) +" hlaS remanescentes")
        else:
            print("Obtendo o resultado do ultimo arquivo excel que tambem tem 10 hlas")
# n_remanescentes = 6
# getEachExcel(n_remanescentes)

# from netMHCpan.Modules.modules import putResultsIntoExcel as putResultsIntoExcel
from NetMHCpan.Captura_do_raw_data.Modules.modules import putResultsIntoExcel as putResultsIntoExcel

global sheet
global hla_list
global epitopo
global excel_path # precisa porque vai salvar no mesmo diretorio

def getMinimoRnaking_sb_wb(sheet, hla_list, epitopo, excel_path):
# def getMinimoRnaking_sb_wb(sheet, epitopo, excel_path):
    coluna_nm = str
    coluna_hla = str
    coluna_ranking = str
    binders = int
    sb = int
    wb = int

    final = False
    linha = 3 # porque é a partir da linha 3 que comeca o ranking

    for i in range(len(hla_list)):

        if (i + 1) == 1:
            coluna_hla = "E"
            coluna_ranking = "F"
            coluna_nm = "E"

        if (i + 1) == 2:
            coluna_hla = "H"
            coluna_ranking = "I"
            coluna_nm = "H"

        if (i + 1) == 3:
            coluna_hla = "K"
            coluna_ranking = "L"
            coluna_nm = "K"

        if (i + 1) == 4:
            coluna_hla = "N"
            coluna_ranking = "O"
            coluna_nm = "N"

        if (i + 1) == 5:
            coluna_hla = "Q"
            coluna_ranking = "R"
            coluna_nm = "Q"

        if (i + 1) == 6:
            coluna_hla = "T"
            coluna_ranking = "U"
            coluna_nm = "T"

        if (i + 1) == 7:
            coluna_hla = "W"
            coluna_ranking = "X"
            coluna_nm = "W"

        if (i + 1) == 8:
            coluna_hla = "Z"
            coluna_ranking = "AA"
            coluna_nm = "Z"

        if (i + 1) == 9:
            coluna_hla = "AC"
            coluna_ranking = "AD"
            coluna_nm = "AC"

        if (i + 1) == 10:
            coluna_hla = "AF"
            coluna_ranking = "AG"
            coluna_nm = "AF"

        print("coluna_hla: ", coluna_hla)
        print("coluna_ranking: ", coluna_ranking )
        print("coluna_nm: ", coluna_nm)

        while not final:
            # print(sheet["H" + str(linha)].value)
            if str(sheet[coluna_ranking + str(linha)].value) != 'None':
                linha = linha + 1
            else:
                linha = (linha - 1)
                print("Ultimo Ranking: ", sheet[coluna_ranking + str(linha)].value)
                print("linha: ", linha)
                final = True
                print("Final")

        # Busca o menor ranking
        print("Buscando o menor ranking")

        #hla
        print("Hla de numero ", i + 1)
        # Inicialmente, o primeiro ranking é o menor
        binders = 0
        sb = 0
        wb = 0

        hla = str(sheet[coluna_hla + "1"].value)
        ranking_minimo = float(sheet[coluna_ranking + "3"].value)
        nm_minimo = float(sheet[coluna_nm + "3"].value)

        for index in range(linha - 4):

            ranking_corrente = float(sheet[coluna_ranking + str(index + 4)].value)
            print("linha: ", index + 4, ".Ranking corrente: ", ranking_corrente)

            if ranking_corrente <= 2:
                binders = binders + 1
                print("binder! binders: ", binders)
                if ranking_corrente < 0.5:
                    sb = sb + 1
                    print("strongBinder! sb = ", sb)

            if ranking_corrente < ranking_minimo:
                ranking_minimo = ranking_corrente
                print(">>>>>>>>>>ranking minimo atualizado para " + str(ranking_corrente))

            nm_corrente = float(sheet[coluna_nm + str(index + 4)].value)

            if nm_corrente < nm_minimo:
                nm_minimo = nm_corrente

        wb = binders - sb
        print("binders, strong binders e weak binders: ")
        print(binders, sb, wb)

        # print("ranking_minimo: ", ranking_minimo)
        # sheet["E" + str(1)] = ranking_minimo
        # print(sheet["E" + str(1)].value)

        print("binders: ", binders)
        print("sb: ", sb)
        print("wb: ", wb)

        results = [hla, ranking_minimo, sb, wb, nm_minimo]
        print("result: ", results)

        putResultsIntoExcel(epitopo, results, excel_path)

# excel_path = "C:\\Users\\Bruno Conde\\Documents\\LGHM\\Automação\\SCRIPTS\\netMHCpan\\results_netMHCpan\\"
# index = 0 # index = numero do excel
# book = load_workbook(excel_path + str(index + 1) + ".xlsx")
# sheet = book.active
# hla_list = ["HLA1", "HLA2", "HLA3", "HLA4", "HLA5", "HLA6", "HLA7", "HLA8", "HLA9", "HLA10"]
# epitopo = "varEpitopos"
#
# getMinimoRnaking_sb_wb(sheet, hla_list, epitopo, excel_path)
