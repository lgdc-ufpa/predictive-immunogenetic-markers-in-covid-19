import os
from openpyxl import load_workbook, Workbook

try:
    book_result = load_workbook(os.getcwd() + '/result.xlsx')

except:
    book_result = Workbook()
    sheet_result = book_result.active
    book_result.save(os.getcwd() + '/result.xlsx')

book_result.close()

hla_directory = os.getcwd() + '/files/'
hla_files = sorted(os.listdir(hla_directory), reverse=False)

for elem in hla_files:
    if '.xls.xlsx' in elem:
        os.rename(str(hla_directory + elem), str(hla_directory + elem[:2] + 'xlsx'))

hla_directory = os.getcwd() + '/files/'
hla_files = sorted(os.listdir(hla_directory), reverse=False)

for index_hla, hla_file in enumerate(hla_files):
    proteins = dict()
    # proteins[protein_name] = [NumBinders, NumStrongBinders]

    hla_book = load_workbook(hla_directory + hla_file)
    hla_sheet = hla_book.active
    hla_name = str(hla_sheet['D1'].value).replace(':', '_')
    print(hla_file, hla_name)

    index = 3
    while str(hla_sheet['C' + str(index)].value) != 'None':

        if hla_sheet['C' + str(index)].value in proteins:

            if hla_sheet['H' + str(index)].value < 2:
                proteins[hla_sheet['C' + str(index)].value][0] += 1

                if hla_sheet['H' + str(index)].value < 0.5:
                    proteins[hla_sheet['C' + str(index)].value][1] += 1
        else:

            proteins[hla_sheet['C' + str(index)].value] = [0, 0]
            # print(type(hla_sheet['H' + str(index)].value), hla_sheet['H' + str(index)].value )
            if hla_sheet['H' + str(index)].value < 2:
                proteins[hla_sheet['C' + str(index)].value][0] += 1

                if hla_sheet['H' + str(index)].value < 0.5:
                    proteins[hla_sheet['C' + str(index)].value][1] += 1

        index += 1

    hla_book.close()

    """
    Salvando a analise do hla
    """

    # book_result = load_workbook(os.getcwd() + '/result_ ' + hla_name + '.xlsx')
    book_result = load_workbook(os.getcwd() + '/result.xlsx')

    book_result.create_sheet(hla_name)
    book_result.save(os.getcwd() + '/result.xlsx')
    sheet_result_hla = book_result[hla_name]

    sheet_result_hla['A1'] = 'Protein'
    sheet_result_hla['B1'] = 'NB'
    sheet_result_hla['C1'] = 'SB'

    for index, protein in enumerate(proteins.keys()):

        nb, sb = proteins[protein][0], proteins[protein][1]
        # print(protein, nb, sb)

        #hla_sheet
        # [proteina, NB, SB]
        sheet_result_hla['A' + str(index + 2)] = protein
        sheet_result_hla['B' + str(index + 2)] = nb
        sheet_result_hla['C' + str(index + 2)] = sb

    book_result.save(os.getcwd() + '/result.xlsx')
    book_result.close()

try:
    book_result = load_workbook(os.getcwd() + '/result.xlsx')
    book_result.remove_sheet(worksheet=book_result['Sheet'])
    book_result.save(os.getcwd() + '/result.xlsx')
    book_result.close()
except:
    pass