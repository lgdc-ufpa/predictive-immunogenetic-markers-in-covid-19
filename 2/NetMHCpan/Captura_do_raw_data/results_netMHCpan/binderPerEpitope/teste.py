import os
from openpyxl import load_workbook, Workbook

try:
    book_result = load_workbook(os.getcwd() + '/result.xlsx')
    print('abriu')

except:
    book_result = Workbook()
    sheet_result = book_result.active
    book_result.save(os.getcwd() + '/result.xlsx')
    print('criou')


# book_result = load_workbook(os.getcwd() + '/result.xlsx')
# sheet_result = book_result.active
#
# for column in sheet_result.iter_cols():
#     for cell in column:
#         if str(cell.value) == 'None':
#             print(column)
#             break

book = load_workbook(os.getcwd() + '/result.xlsx')
book.remove_sheet(worksheet=book['nova'])
book.create_sheet('nova')
book.save(os.getcwd() + '/result.xlsx')
book.close()