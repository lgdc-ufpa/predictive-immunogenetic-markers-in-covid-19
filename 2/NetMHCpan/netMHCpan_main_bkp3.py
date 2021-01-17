from selenium import webdriver
from openpyxl import load_workbook
from NetMHCpan.Modules import SalvarEstado
import os

diretorio = "C:\\Users\\mauro\\Desktop\\Bruno\\Automação\\SCRIPTS\\NetMHCpan\\"
driver_path = "C:\\Users\\mauro\\Desktop\\Bruno\\Automação\\configuring the envyronmet\\chromedriver\\chromedriver_win32\\chromedriver.exe"

#variaveis
ultima_analise = 0
numero_analises = 0
n_alelos_remanescentes = 0

#estado
epitopos_txt = str
epitopos_path = str
alelos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\alelostestefinal.xlsx"
epitopos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\Epitopos\\"

abrir = False
while not abrir:
    try:
        nome_epitopos = str(input("Entre com o nome do epitopo sem a extensão .txt:"))
        epitopos_txt = open(epitopos_path + nome_epitopos+'.txt', "r")
        epitopos = (epitopos_txt.readlines())
        abrir = True
    except:
        print("Nome incorreto. Insira o nome sem a extensão .txt")

##########
book_alelos = load_workbook(alelos_path)
sheet = book_alelos.active

alelos = []
numero_analises = 0
n_alelos_remanescentes = int
linha = int

final = bool

# checa se há registros
try:
    file = open(diretorio + nome_epitopos + '_estado.txt', 'r')
    estados = file.readlines()
    print(estados)
    file.close()
    ultima_analise = int(estados[0])
    numero_analises = int(estados[1])
    n_alelos_remanescentes = int(estados[2])
    print("Existe um estado")
    print("Ultima analise: ", ultima_analise, type(ultima_analise))

    if ultima_analise == int(estados[1]):
        print("Analise Final")
        final = True
    else:
        final = False

except:

    final = False
    linha = 0
    while (not final):
        linha = linha + 1
        # print(sheet["A"+str(linha)].value)
        if str(sheet["A" + str(linha)].value) == "None":
            print("Acabou")
            final = True
            linha = linha - 1
    final = False
    print(linha)
    print("numero de analises de 10: ", int(linha / 10))
    print("numero de alelos da ultima analise: ", linha % 10)

    if (linha % 10 == 0):
        numero_analises = int(linha / 10)
    else:
        numero_analises = int(linha / 10) + 1
        n_alelos_remanescentes = linha % 10

    SalvarEstado.salvarEstado(diretorio, nome_epitopos, ultima_analise, numero_analises, n_alelos_remanescentes)


##########

while(ultima_analise != numero_analises):
    driver = webdriver.Chrome(driver_path)

    driver.get("http://www.cbs.dtu.dk/services/NetMHCpan/")
    textarea = driver.find_element_by_tag_name('textarea')
    textarea.send_keys(epitopos)

    alelos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\alelostestefinal.xlsx"
    book_alelos = load_workbook(alelos_path)
    sheet = book_alelos.active

    alelos = []

    if (not final):
        for i in range(10):
            print(i+1, sheet["A"+str((i*(ultima_analise+1)) + 1)].value)
            if not i == 9:
                alelos.append(sheet["A"+str((i*(ultima_analise+1)) + 1)].value + ",")
            else:
                alelos.append(sheet["A"+str((i*(ultima_analise+1)) + 1)].value)
        print(alelos)
    else:
        for i in range(n_alelos_remanescentes):
            print(i+1, sheet["A"+str((i*(ultima_analise+1)) + 1)].value)
            if not i == 9:
                alelos.append(sheet["A"+str((i*(ultima_analise+1)) + 1)].value + ",")
            else:
                alelos.append(sheet["A"+str((i*(ultima_analise+1)) + 1)].value)
        print(alelos)

    alele_area = driver.find_element_by_name("allele")
    alele_area.send_keys(alelos)

    driver.find_element_by_name("BApred").click()
    driver.find_element_by_name("sort").click()
    driver.find_element_by_name("xlsdump").click()

    #Submit
    inputList = driver.find_elements_by_tag_name("input")
    inputList[-2].click()

    #Output
    link_list = driver.find_elements_by_tag_name("a")
    link_list[-3].click()

    ultima_analise = ultima_analise + 1
    SalvarEstado.salvarEstado(diretorio, nome_epitopos, numero_analises, ultima_analise, n_alelos_remanescentes)
else:

    os.remove(diretorio + nome_epitopos + '_estado.txt')

print("##############Finalizado################")









