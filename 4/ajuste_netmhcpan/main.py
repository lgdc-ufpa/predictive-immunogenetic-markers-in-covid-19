from selenium import webdriver
from openpyxl import load_workbook
# from Modules import SalvarEstado
from time import sleep
import os
import pyautogui
import clipboard

# diretorio = "C:\\Users\\mauro\\OneDrive\\Doutorado\\Automação\\scripts"
diretorio = os.getcwd()
# driver_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\Automação\\scripts\\chromedriver.exe"
driver_path = diretorio + '\\geckodriver'

#variaveis
# ultima_analise = 1
# numero_analises = 0
# n_alelos_remanescentes = 0

#MUDANCA ----------- PARA TESTE APENAS
# alelos = []
alelos = "HLA-A01:01"
linha = int
final = bool

# alelos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\alelostestefinal.xlsx"
alelos_path = diretorio + '\\alelostestefinal.xlsx'

book_alelos = load_workbook(alelos_path)
sheet = book_alelos.active

epitopos_path = str
epitopos_txt = str
# epitopos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\Epitopos\\"
epitopos_path = diretorio + '\\'
epitopos = ""

#Tenta abrir o arquivo dos epitopos
abrir = False
while not abrir:
    try:
        nome_epitopos = str(input("Entre com o nome do epitopo sem a extensão .txt:"))
        epitopos_txt = open(epitopos_path + nome_epitopos+'.txt', "r")
        for linha in epitopos_txt:
            if not epitopos == "":
                epitopos += '\n' + linha.replace('\n', "")
            else:
                epitopos = linha

        abrir = True

    except:
        print("Nome incorreto. Insira o nome sem a extensão .txt")


# # checa se há registros
# try:
#     file = open(diretorio + nome_epitopos + '_estado.txt', 'r')
#     estados = file.readlines()
#     print(estados)
#     file.close() #TODO: testar estados[0], [1] e [2] e os tipos
#     ultima_analise = int(estados[0])
#     numero_analises = int(estados[1])
#     n_alelos_remanescentes = int(estados[2]) #TODO: testar essas 3 variaveis do estado
#     print("Existe um estado")
#     print("ultima analise: ", ultima_analise, type(ultima_analise))
#     print("numero_analises: ", numero_analises, type(numero_analises))
#     print("n_alelos_remanescentes: ", n_alelos_remanescentes, type(n_alelos_remanescentes))
#
#     # checa se a ultima_analise feita é a análise final
#     if ultima_analise == int(estados[1]):
#         print("Analise Final")
#         final = True
#     else:
#         final = False
#
# except:
#     print("Não há estados")
#     final = False
#     linha = 0
#     while (not final):
#         linha = linha + 1
#         print(sheet["A"+str(linha)].value)
#         if str(sheet["A" + str(linha)].value) == "None":
#             print("Acabou")
#             final = True
#             linha = linha - 1
#     final = False
#     print("numero de alelos: ", linha)
#     print("numero de analises de 1: ", int(linha / 1))
#     print("numero de alelos da ultima analise: ", linha % 1)
#
#     #checa se todas as análises serão de 10 alelos
#     if (linha % 2 == 0):
#         numero_analises = int(linha / 1)
#
#     #acrescenta uma analise a mais para os alelos remanescentes
#     else:
#         numero_analises = int(linha / 1) + 1
#         n_alelos_remanescentes = linha % 1
#
#     #cria o estado
#     # ultima analise e numeroamalises estao com valores trocados
#     file = open(diretorio + nome_epitopos + '_estado.txt', 'w')
#     file.write(str(ultima_analise) + '\n' + str(numero_analises) + '\n' + str(n_alelos_remanescentes))
#     file.close()
#
#     # checa se há registros
#     file = open(diretorio + nome_epitopos + '_estado.txt', 'r')
#     estados = file.readlines()
#     print(estados)
#     estado = file.readline()
#     print(estado)
#     file.close()

#netMHCpan
print("Entra no netMHCpan e coloca os epitopos")
driver = webdriver.Firefox()
# driver.get("https://services.healthtech.dtu.dk/services/NetMHCpan-4.0/1-Submission.php")
driver.get("https://services.healthtech.dtu.dk/service.php?NetMHCpan-4.0")


#MUDANCA#---------------------
# driver.find_element_by_name("SEQPASTE").click()
# driver.find_element_by_tag_name('textarea').click()
pyautogui.press('pagedown')
sleep(0.3)
pyautogui.leftClick(400,400)
sleep(0.3)

#É aqui
clipboard.copy(epitopos)
pyautogui.hotkey('ctrl', 'v')
sleep(0.3)
pyautogui.press('tab')
sleep(0.3)
pyautogui.press('tab')
sleep(0.3)
pyautogui.press('tab')
sleep(0.3)
pyautogui.press('down')
pyautogui.press('down')

#Xpath = '//option[@value="9"]'
#option_9mer = driver.find_element_by_xpath(Xpath)
#print(option_9mer.text)
#option_9mer.click()
#sleep(20)

print("pega os alelos")
# alelos_path = "C:\\Users\\mauro\\OneDrive\\Doutorado\\alelostestefinal.xlsx"
alelos_path = diretorio + '\\alelostestefinal.xlsx'
book_alelos = load_workbook(alelos_path)
sheet = book_alelos.active
print("Reseta a variável dos alelos para pegar os alelos da próxima análise")
alelos = []
print("Insere os seguintes alelos:")

#TIRAR ISSO DEPOIS, POIS E SO PARA TESTE SEM A RECUPERACAO DE ESTADO
ultima_analise = 0

if (not final):
    print("insere os 1")
    for i in range(1):
        print(i+1, sheet["A"+str(((ultima_analise - 1)) + 1 + i)].value)
        if not i == 1:
            alelos.append(sheet["A"+str(((ultima_analise - 1)) + 1 + i)].value)
        else:
            alelos.append(sheet["A"+str(((ultima_analise - 1)) + 1 + i)].value)
    print(alelos)

#netMHCpan
print("netMHCpan: coloca os alelos")

#MUDANCA -----------------------------------
# alele_area = driver.find_element_by_name("allele")
# alele_area.send_keys(alelos)
pyautogui.leftClick(0,400)
sleep(1.0)
pyautogui.press('pagedown')
sleep(0.3)
pyautogui.leftClick(400,275)
#it work only for one alele, cause a list of aleles should be concatenated, separated by commmas, and no space[]
clipboard.copy(str(alelos))
pyautogui.hotkey('ctrl','v')
sleep(0.3)
pyautogui.leftClick(500,290)
sleep(0.3)
pyautogui.press('pagedown')
sleep(0.3)
pyautogui.leftClick(1250,500)
sleep(0.3)
pyautogui.press('pageup')
sleep(0.3)
#make ba predictions
pyautogui.leftClick(380,425)
sleep(0.3)
#sort by prediction score
pyautogui.leftClick(395,460)
sleep(0.3)
#save predictions to xls file
pyautogui.leftClick(420,500)
sleep(0.3)
#submit
pyautogui.leftClick(230,540)
sleep(0.3)

#MUDANCA ------- ESTES METODOS FORAM SUBSTITUIDOS PELOS METODOS ACIMA
# driver.find_element_by_name("BApred").click()
# driver.find_element_by_name("sort").click()
# driver.find_element_by_name("xlsdump").click()

print("netMHCpan: submete p/ analise e gera a saída")
# Submit
inputList = driver.find_elements_by_tag_name("input")
inputList[-2].click()

print("nethMHCpan: gera o excel de saida")
# Output
clicou = False
esperando = True
while (not clicou):
    try:
        driver.find_elements_by_tag_name("a")[-3].click()
        clicou = True
    except:
        if esperando:
            print("Esperando carregar")
            esperando = False

print('\nFIM')